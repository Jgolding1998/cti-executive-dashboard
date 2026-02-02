import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Style definitions
header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
section_fill = PatternFill(start_color='3D5A80', end_color='3D5A80', fill_type='solid')
section_font = Font(bold=True, color='FFFFFF', size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_section_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border

def add_data_row(ws, row, data):
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border

# Remove default sheet
del wb['Sheet']

# ============ OVERVIEW PAGE ============
ws = wb.create_sheet('1. Overview Page')
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

row = 1
ws.cell(row=row, column=1, value='CTI EXECUTIVE DASHBOARD - DATA MAPPING')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1).font = Font(bold=True, size=14)
row += 2

# Header
headers = ['Component', 'Data Source / IDO', 'Fields Used', 'Calculation / Logic']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

# KPI Cards section
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='KPI CARDS (Top Row)')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ["Yesterday's Sales", 'SLLedgers (GL Ledger)', 'Acct (401000, 402000, 495000, 495400), DomAmount, TransDate, Ref', 'Filter TransDate = last business day (skip weekends). Sum -DomAmount for revenue accounts. Credits are negative in GL so multiply by -1.'],
    ['Month-to-Date', 'SLLedgers (GL Ledger)', 'Same as above', 'Filter TransDate >= first of current month. Sum all daily totals.'],
    ['Year-to-Date', 'SLLedgers (GL Ledger)', 'Same as above', 'Filter TransDate >= Jan 1 of current year. Sum all daily totals.'],
    ['AR Outstanding', 'Birst AR Aging (birst-ar-aging.json) OR SLArTrans', 'Current, Days1_30, Days31_60, Days61_90, Days90Plus', 'Sum all aging buckets. Birst data is source of truth if available, otherwise calculated from AR transactions.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='60-DAY SALES TREND CHART')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Daily Trend Lines', 'SLLedgers (GL Ledger)', 'Acct, DomAmount, TransDate, Ref', 'Loop 60 days back from today (weekdays only). For each day, aggregate GL entries by account. Product/Service split uses invoice-level categorization.'],
    ['Product vs Service Split', 'SLCoItems + SLItems', 'Item, ProductCode, DerExtInvoicedPrice', 'Build order breakdown: LFTR/LGAS/LIHL ProductCodes = Service. LMSC/XNIO = Misc. Everything else = Product. Apply ratio to GL amount per invoice.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value="YESTERDAY'S BREAKDOWN")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Total', 'SLLedgers', 'DomAmount for Acct 401000, 402000, 495000, 495400', 'Sum of all revenue account entries for yesterday * -1'],
    ['Product Amount', 'SLLedgers + SLCoItems + SLItems', 'Via invoice reference (Ref field)', 'GL amount * (order product total / order total) using ProductCode categorization'],
    ['Service Amount', 'SLLedgers + SLCoItems + SLItems', 'Via invoice reference', 'GL amount * (order service total / order total). Service = ProductCode in (LFTR, LGAS, LIHL)'],
    ['Freight Amount', 'SLLedgers', 'Acct = 495400', 'Direct from GL account 495400 (Freight Revenue)'],
    ['Miscellaneous Amount', 'SLLedgers', 'Acct = 495000 OR ProductCode in (LMSC, XNIO)', 'Account 495000 direct + ProductCode misc items'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='MTD SALES MIX (Pie Chart)')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Category Percentages', 'Derived from MTD totals', 'Product, Service, Freight, Miscellaneous', 'Each category / MTD Total * 100. Display as pie chart segments.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='DAILY AVERAGES')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['MTD Daily Average', 'Calculated', 'MTD Total, Business Days', 'MTD Total / MTD Business Days. Business days exclude weekends and US federal holidays.'],
    ['YTD Daily Average', 'Calculated', 'YTD Total, Business Days', 'YTD Total / YTD Business Days'],
    ['MTD Business Days', 'Calculated', 'Date range', 'Count weekdays from month start to today, excluding holidays list'],
    ['YTD Business Days', 'Calculated', 'Date range', 'Count weekdays from Jan 1 to today, excluding holidays list'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='AR AGING')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Current', 'Birst OR SLArTrans', 'DueDate, Amount, Type, ApplyToInvNum', 'Invoices where DueDate >= today. Balance = Original - Payments - Credits'],
    ['1-30 Days', 'Birst OR SLArTrans', 'Same', 'Invoices where DaysOverdue 1-30'],
    ['31-60 Days', 'Birst OR SLArTrans', 'Same', 'Invoices where DaysOverdue 31-60'],
    ['61-90 Days', 'Birst OR SLArTrans', 'Same', 'Invoices where DaysOverdue 61-90'],
    ['90+ Days', 'Birst OR SLArTrans', 'Same', 'Invoices where DaysOverdue > 90'],
    ['Balance Calculation', 'SLArTrans', 'Type (I=Invoice, P=Payment, C=Credit)', 'For each invoice: Sum Type=I amounts, subtract Type=P and Type=C where ApplyToInvNum matches'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='DAY OF WEEK PERFORMANCE')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['DOW Averages', 'SLLedgers (60-day history)', 'TransDate, DomAmount', 'Group last 60 days by day of week. Average = Sum of DOW totals / Count of that DOW. If no data, default to MTD average.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='TOP CUSTOMERS MTD')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Customer Sales', 'SLArTrans + SLLedgers', 'CadName (customer name), InvNum, CoNum', 'Link GL entry Ref (e.g. "ARI 12345") to AR invoice, get customer name. Aggregate by customer, sort descending, take top 15.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='TOP PRODUCTS MTD')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Product Sales', 'SLCoItems', 'Item, ItDescription, DerExtInvoicedPrice', 'Aggregate DerExtInvoicedPrice by Item. Sort descending, take top 15. Category from ProductCode lookup.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Overview page created')

# ============ SALES TEAM PAGE ============
ws = wb.create_sheet('2. Sales Team Page')
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

row = 1
headers = ['Component', 'Data Source / IDO', 'Fields Used', 'Calculation / Logic']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='TEAM KPI CARDS')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Yesterday Total (All)', 'Derived from salesperson aggregation', 'Same as Overview sales', 'Sum of all team member yesterday sales'],
    ['MTD Total (All)', 'Derived from salesperson aggregation', 'Same as Overview sales', 'Sum of all team member MTD sales'],
    ['Top Performer MTD', 'Calculated', 'Salesperson MTD amounts', 'Team member with highest MTD total'],
    ['MTD Invoices', 'SLCoS + SLLedgers', 'CoNum linked to GL via invoice', 'Count of unique invoices per salesperson for MTD'],
    ['Service Revenue MTD', 'Calculated', 'MTDService from salesperson records', 'Sum of service amounts for service reps'],
    ['Avg Service Call', 'Calculated', 'Service revenue / invoice count', 'Total service revenue / number of service invoices'],
    ['States Covered MTD', 'SLCoS', 'ShipToState', 'Count distinct states from service rep orders'],
    ['Top Service Rep', 'Calculated', 'Service rep MTD amounts', 'Service rep with highest MTD total'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='SALESPERSON TRACKING')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Order to Salesperson Map', 'SLCoS (Customer Orders)', 'CoNum, TakenBy, OrderDate, Invoiced, ShipToCity, ShipToState', 'Filter: Invoiced=1, OrderDate >= YearStart. Build lookup: orderToSalesperson[CoNum] = TakenBy'],
    ['Invoice to Order Map', 'SLArTrans', 'InvNum, CoNum', 'Build lookup: invToOrder[InvNum] = CoNum'],
    ['GL to Salesperson', 'Chained lookup', 'GL Ref -> InvNum -> CoNum -> TakenBy', 'Parse GL Ref for invoice number (regex "ARI (\\d+)"), lookup order, lookup salesperson'],
    ['Role Assignment', 'Hardcoded teamRoles dictionary', 'Name -> Role, Type', 'Roles: COO, Accounting, Sales Rep, Service Rep. Types: Executive, Admin, Sales, Service'],
    ['Territory Tracking', 'SLCoS', 'ShipToState, ShipToCity per TakenBy', 'Aggregate distinct states and cities per salesperson from their orders'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='PRODUCT/SERVICE SPLIT BY PERSON')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['MTD Product Amount', 'Calculated per salesperson', 'GL amount * product ratio', 'txnProdAmt = GL amount * (order product total / order total)'],
    ['MTD Service Amount', 'Calculated per salesperson', 'GL amount * service ratio', 'txnSvcAmt = GL amount * (order service total / order total)'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Sales Team page created')

# ============ CASH FLOW PAGE ============
ws = wb.create_sheet('3. Cash Flow Page')
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 60

row = 1
headers = ['Component', 'Data Source / IDO', 'Fields Used', 'Calculation / Logic']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='CASH FLOW KPI CARDS')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['This Week Forecast', 'Calculated', 'Week 1 of cashFlowPrediction', 'Actual (past days) + Predicted (future days) for current week'],
    ['Next Week Forecast', 'Calculated', 'Week 2 of cashFlowPrediction', 'Fully predicted based on algorithm'],
    ['6-Week Total', 'Calculated', 'Sum all 6 weeks', 'Sum of weekly totals'],
    ['AR Due Next 6 Weeks', 'SLArTrans', 'DueDate, Amount (open invoices)', 'Sum of open AR with DueDate in next 6 weeks'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='PREDICTION ALGORITHM')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Historical Daily Average', 'SLLedgers', 'MTD sales total / MTD business days', 'Baseline = MTD average daily sales'],
    ['AR Collection Rates', 'Defined constants', 'Per-bucket weekly rates', 'Current: 20%/wk, 1-30d: 15%/wk, 31-60d: 10%/wk, 61-90d: 5%/wk, 90+: 2%/wk'],
    ['Day of Week Weights', 'Defined constants', 'Distribution across weekdays', 'Mon: 15%, Tue: 25%, Wed: 25%, Thu: 20%, Fri: 15%'],
    ['Week Expected Collections', 'Calculated', 'Blend of historical + AR', 'weekExpectedCollections = (historicalDailyAvg * 5 * 0.6) + (arBasedExpectation * 0.4)'],
    ['Daily Prediction', 'Calculated', 'DOW avg + week spread', 'dayPrediction = (DOW_Avg * 0.7 + weekExpected * DOW_weight * 0.3) * holidayFactor'],
    ['Holiday Factor', 'Holiday list', 'Date matching', 'If date in holidays list: factor = 0.2, else factor = 1.0'],
    ['AR Aging Roll', 'Calculated per week', 'Bucket transitions', 'Each bucket ages: 25% moves to next bucket, collections reduce balance'],
    ['Confidence Score', 'Calculated', 'Distance + actual %', 'Base: Wk1=90%, Wk2=75%, Wk3-4=60%, Wk5-6=45%. Adjusted by actual data proportion.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 1
style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='ACTUAL VS PREDICTED')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Actual (past days)', 'SLLedgers', 'TransDate < today', 'Use actual GL data for days already passed'],
    ['Today', 'SLLedgers', 'TransDate = today', 'Use actual if posted, otherwise $0'],
    ['Predicted (future)', 'Algorithm', 'TransDate > today', 'Apply prediction formula'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Cash Flow page created')

# ============ CUSTOMERS PAGE ============
ws = wb.create_sheet('4. Customers Page')
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

row = 1
headers = ['Component', 'Data Source / IDO', 'Fields Used', 'Calculation / Logic']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='CUSTOMER DATA')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Top 15 Customers MTD', 'SLArTrans + SLLedgers', 'CadName, InvNum, linked via GL Ref', 'Aggregate sales by customer for MTD period. Sort desc, take top 15. Display as bar chart.'],
    ["Yesterday's Top Customers", 'SLArTrans + SLLedgers', 'Same, filtered by date', 'Same aggregation filtered to yesterday only. Sort desc, take top 15.'],
    ['Full Customer List MTD', 'SLArTrans + SLLedgers', 'Same', 'All customers with MTD sales > 0, sorted by amount descending'],
    ['Customer Name Source', 'SLArTrans', 'CadName field', 'Customer name from AR transaction record (invoice header)'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Customers page created')

# ============ PRODUCTS PAGE ============
ws = wb.create_sheet('5. Products Page')
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

row = 1
headers = ['Component', 'Data Source / IDO', 'Fields Used', 'Calculation / Logic']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='PRODUCT DATA')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Top 15 Products MTD', 'SLCoItems', 'Item, ItDescription, DerExtInvoicedPrice', 'Aggregate DerExtInvoicedPrice by Item. Sort desc, take top 15.'],
    ['Product Category Mix', 'SLCoItems + SLItems', 'ProductCode for categorization', 'Group products by category (Product/Service/Misc). Show as pie chart.'],
    ['Full Product List', 'SLCoItems', 'All with MTD sales > 0', 'Full list sorted by revenue descending'],
    ['Category Assignment', 'SLItems', 'ProductCode field', 'LFTR/LGAS/LIHL = Service, LMSC/XNIO = Misc, others = Product'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Products page created')

# ============ GEOGRAPHY PAGE ============
ws = wb.create_sheet('6. Geography Page')
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 50

row = 1
headers = ['Component', 'Data Source / IDO', 'Fields Used', 'Calculation / Logic']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

style_section_row(ws, row, 4)
ws.cell(row=row, column=1, value='SHIPPING/GEOGRAPHY DATA')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

data = [
    ['Shipping Heat Map', 'SLShipments', 'ConsigneeCity, ConsigneeState, ConsigneeZip', 'Aggregate shipments by city/state. Use Leaflet.js with city coordinates for heat map markers.'],
    ['Top Shipping Locations', 'SLShipments', 'Same, with counts', 'Sort by shipment count descending, take top 100'],
    ['Shipments by State', 'SLShipments', 'ConsigneeState', 'Aggregate count by state. Display as horizontal bar chart.'],
    ['Service Tech Territory', 'SLCoS (orders)', 'ShipToState, ShipToCity per TakenBy', 'For service reps (Type=Service), show distinct states/cities covered. Different color per rep on map.'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Geography page created')

# ============ IDO REFERENCE ============
ws = wb.create_sheet('7. IDO Reference')
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 40

row = 1
headers = ['IDO Name', 'SyteLine Table', 'Key Fields Used', 'Purpose in Dashboard']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

data = [
    ['SLLedgers', 'ledger', 'Acct, DomAmount, TransDate, Ref, ControlYear', 'SOURCE OF TRUTH for all sales data. Revenue accounts: 401000 (Sales), 402000 (Sales-Services), 495000 (Misc Revenue), 495400 (Freight Revenue)'],
    ['SLArTrans', 'artrans', 'InvNum, CoNum, CadName, InvDate, DueDate, Amount, Type, ApplyToInvNum', 'AR aging, invoice-to-order mapping, customer names. Type: I=Invoice, P=Payment, C=Credit'],
    ['SLCoItems', 'coitem', 'CoNum, Item, ItDescription, DerExtInvoicedPrice', 'Line item details for product/service categorization. Links order to items.'],
    ['SLItems', 'item', 'Item, ProductCode', 'Product categorization. ProductCode determines Service vs Product vs Misc classification.'],
    ['SLCoS', 'co', 'CoNum, TakenBy, Price, OrderDate, Invoiced, ShipToCity, ShipToState', 'Order header. TakenBy = salesperson. Used for sales attribution and territory tracking.'],
    ['SLShipments', 'shipment', 'ConsigneeCity, ConsigneeState, ConsigneeZip', 'Shipping destination data for geography analysis'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 2
ws.cell(row=row, column=1, value='ACCOUNT STRUCTURE')
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1

headers = ['Account', 'Description', 'Notes', 'Treatment']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

data = [
    ['401000', 'Sales (consolidated)', 'All product sales revenue', 'Main sales account - requires invoice-level breakdown for Product vs Service'],
    ['402000', 'Sales - Services', 'Service sales revenue', 'Some service revenue posts here directly'],
    ['495000', 'Miscellaneous Revenue', 'Misc income', 'Non-standard revenue items'],
    ['495400', 'Freight Revenue', 'Shipping charges', 'Freight billed to customers'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 2
ws.cell(row=row, column=1, value='PRODUCT CODES')
ws.cell(row=row, column=1).font = Font(bold=True)
row += 1

headers = ['Code', 'Description', 'Use', 'Category']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 4)
row += 1

data = [
    ['LFTR', 'Service - Field Labor', 'Field technician labor', 'Counted as Service revenue'],
    ['LGAS', 'Service - Gas', 'Gas detection service', 'Counted as Service revenue'],
    ['LIHL', 'Service - In-house Labor', 'In-house labor charges', 'Counted as Service revenue'],
    ['LMSC', 'Service - Misc', 'Miscellaneous service', 'Counted as Miscellaneous'],
    ['XNIO', 'General Miscellaneous', 'Non-inventory items', 'Counted as Miscellaneous'],
    ['(All others)', 'Product', 'Physical products', 'Counted as Product revenue'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('IDO Reference page created')

# ============ DATA FLOW ============
ws = wb.create_sheet('8. Data Flow')
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 50

row = 1
ws.cell(row=row, column=1, value='DATA FLOW SEQUENCE')
ws.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 2

headers = ['Step', 'Action', 'Output']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 3)
row += 1

data = [
    ['1', 'Fetch SLItems for ProductCode lookup', 'itemProductCode[Item] = ProductCode dictionary'],
    ['2', 'Fetch SLCoItems for line item details', 'orderBreakdown[CoNum] = {Service, Product, Misc, Total} per order'],
    ['3', 'Fetch SLArTrans for invoices', 'invToOrder[InvNum] = CoNum, invToCustomer[InvNum] = CadName, invoiceBalances for AR'],
    ['4', 'Fetch SLCoS for order headers', 'orderToSalesperson[CoNum] = TakenBy, territoryByPerson[TakenBy] = {States, Cities}'],
    ['5', 'Fetch SLLedgers for GL data', 'salesByDate[YYYY-MM-DD] = {Product, Service, Freight, Misc, Total}'],
    ['6', 'Process GL entries', 'For each GL entry: parse Ref for invoice, lookup order, apply category ratio, aggregate by date/customer/salesperson'],
    ['7', 'Calculate summaries', 'Yesterday, MTD, YTD totals and averages'],
    ['8', 'Load/Calculate AR aging', 'Birst data if available, otherwise calculate from invoiceBalances'],
    ['9', 'Run cash flow prediction', '6-week forecast with day-by-day breakdown'],
    ['10', 'Fetch SLShipments', 'Geography heat map data'],
    ['11', 'Generate JSON + HTML', 'dashboard-data.json, cti-dashboard-live.html'],
    ['12', 'Push to GitHub', 'Live at https://jgolding1998.github.io/cti-executive-dashboard/'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

row += 2
style_section_row(ws, row, 3)
ws.cell(row=row, column=1, value='KEY LINKING LOGIC')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
row += 1

data = [
    ['GL -> Invoice', 'Parse GL Ref field using regex "ARI (\\d+)" to extract invoice number', 'Required to link GL amounts back to orders'],
    ['Invoice -> Order', 'Lookup invToOrder[InvNum] from SLArTrans', 'Links invoice to customer order (CoNum)'],
    ['Order -> Salesperson', 'Lookup orderToSalesperson[CoNum] from SLCoS', 'TakenBy field identifies who entered the order'],
    ['Order -> Category', 'Lookup orderBreakdown[CoNum] from SLCoItems + SLItems', 'ProductCode determines Service/Product/Misc split'],
    ['Invoice -> Customer', 'Lookup invToCustomer[InvNum] from SLArTrans', 'CadName field provides customer name'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Data Flow page created')

# ============ FORMULAS REFERENCE ============
ws = wb.create_sheet('9. Key Formulas')
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 40

row = 1
ws.cell(row=row, column=1, value='KEY FORMULAS AND CALCULATIONS')
ws.cell(row=row, column=1).font = Font(bold=True, size=12)
row += 2

headers = ['Calculation', 'Formula', 'Notes']
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, 3)
row += 1

data = [
    ['Revenue from GL', 'Amount = DomAmount * -1', 'GL credits are negative, so multiply by -1 to show positive revenue'],
    ['Service Ratio', 'svcRatio = orderBreakdown[CoNum].Service / orderBreakdown[CoNum].Total', 'Proportion of order that is service items'],
    ['Product Ratio', 'prodRatio = orderBreakdown[CoNum].Product / orderBreakdown[CoNum].Total', 'Proportion of order that is product items'],
    ['Daily Average', 'dailyAvg = totalSales / businessDays', 'Excludes weekends and holidays'],
    ['Business Days', 'Count weekdays where DayOfWeek not in (Sat, Sun) AND date not in Holidays', 'US federal holidays hardcoded'],
    ['AR Balance', 'balance = Sum(Type=I) - Sum(Type=P) - Sum(Type=C) where ApplyToInvNum matches', 'Net of payments and credits'],
    ['Days Overdue', 'daysOverdue = (today - DueDate).Days', 'Negative means not yet due (Current)'],
    ['Cash Flow Weekly', 'weekExpected = (histAvg * 5 * 0.6) + (arExpected * 0.4)', 'Blend of historical pattern and AR expectations'],
    ['Cash Flow Daily', 'dayPrediction = (DOW_Avg * 0.7 + weekExpected * DOW_weight * 0.3) * holidayFactor', 'DOW pattern with week normalization'],
    ['Confidence', 'confidence = baseConfidence * (0.5 + actualPct * 0.5)', 'Higher when more actual data available'],
]
for d in data:
    add_data_row(ws, row, d)
    row += 1

print('Formulas page created')

# Save
output_path = r'C:\Users\justi\clawd\dashboard\CTI_Dashboard_Data_Mapping.xlsx'
wb.save(output_path)
print(f'\nExcel file saved successfully to: {output_path}')
